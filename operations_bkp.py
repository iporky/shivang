import spacy
import pandas as pd
import numpy as np
import mkit
from spacy.matcher import Matcher
from spacy.tokens import Doc
from PyPDF2 import PdfFileReader
import slate3k as slate
import os
from openpyxl import load_workbook

df = pd.read_excel('/data/Helpdesk_Report_part_cleaned_final.xlsx')

unprocessDocs = []
docIndex = {'id': 1}
sampleDocs = []
for index in df.index:
    # finalNotes = ''
    # notes = str(df["Notes"][index])
    # notesSplit = notes.split(";")
    # cleanedArray = []
    # if notesSplit[0] != 'nan':
    #   for note in notesSplit:
    #       if note:
    #         finalNotes+= '. ' + note
    #         cleanedArray.append(note)
    text = str(df["Ticket Description"][index])
    context = {"closure_note": str(df["Ticket Closure notes"][index]), "desc": str(df["Ticket Description"][index])}#, "note": finalNotes, "noteArray": cleanedArray}
    unprocessDocs.append((text, context))

nlp = spacy.load('en_vectors_web_lg')
category_list = [nlp('error'),nlp('access'),nlp('usage'),nlp('report'),nlp('logs')]
matcher = Matcher(nlp.vocab)

appMap = {
    "data distribution": "dd",
    "nonconfirmance management system": "nms",
    "significant process substantiation": "sps",
    "high performance archive system": "hpas",
    "quality escape management": "qem",
    "engineering program execution tool": "epex",
    "design reviews": "dr",
    "design review": "dr",
    "design practice": "dp",
    "drawing signature record": "dsr",
    "design record book": "drb",
    "repairbuilder": "reb",
    "repair builder": "reb",
    "export tagging": "dax","nginxutil": "nginx","openresty":"nginx",
    "fai": "cav","cnf": "cav", "cid": "ecid", "aeip":"afrt","tc": "teamcenter",
    "digital thread engineer": "dte", "oneidm": "idm", "data lake": "datalake"
}

matcher.add("Application", None,
            [{"LOWER": "sps"}],[{"LOWER": "drb"}],[{"LOWER": "nms"}],[{"LOWER": "cert"}],[{"LOWER": "dwb"}],
            [{"LOWER": "dd"}],[{"LOWER": "dp"}],[{"LOWER": "dr"}],[{"LOWER": "dsr"}],
            [{"LOWER": "epex"}],[{"LOWER": "hpas"}],[{"LOWER": "qem"}],[{"LOWER": "dti"}],
            [{"LOWER": "dte"},{"LOWER": "discover"}],[{"LOWER": "departure"},{"LOWER": "record"}],
            [{"LOWER": "dp"},{"LOWER": "systems"}],[{"LOWER": "dr"},{"LOWER": "systems"}],
            [{"LOWER": "drb"},{"LOWER": "systems"}],[{"LOWER": "laws"},{"LOWER": "fa"}],
            [{"LOWER": "repair"},{"LOWER": "docs"}],[{"LOWER": "aapi"}],[{"LOWER": "dax"}],
            [{"LOWER": "significant"},{"LOWER": "process"},{"LOWER": "substantiation"}],
            [{"LOWER": "design"},{"LOWER": "record"},{"LOWER": "book"}],
            [{"LOWER": "nonconfirmance "},{"LOWER": "management"},{"LOWER": "system"}],
            [{"LOWER": "data"},{"LOWER": "distribution"}],[{"LOWER": "design"},{"LOWER": "practice"}],
            [{"LOWER": "design"},{"LOWER": "reviews"}],[{"LOWER": "ppm"}],
            [{"LOWER": "drawing"},{"LOWER": "signature"},{"LOWER": "record"}],
            [{"LOWER": "engineering"},{"LOWER": "program"},{"LOWER": "execution"},{"LOWER": "tool"}],
            [{"LOWER": "high"},{"LOWER": "performance"},{"LOWER": "archive"},{"LOWER": "system"}],
            [{"LOWER": "quality"},{"LOWER": "escape"},{"LOWER": "management"}],[{"LOWER": "teamcenter"}],[{"LOWER": "tc"}],
            [{"LOWER": "reb"}],[{"LOWER": "repair"},{"LOWER": "builder"}],[{"LOWER": "repairbuilder"}],
            [{"LOWER": "export"}, {"LOWER": "tagging"}],[{"LOWER": "cav"}],[{"LOWER": "cnf"}],[{"LOWER": "fai"}],
            [{"LOWER": "nginx"}], [{"LOWER": "nginxutil"}], [{"LOWER": "kibana"}], [{"LOWER": "jenkins"}],[{"LOWER": "ecid"}],
            [{"LOWER": "cid"}], [{"LOWER": "dte"}], [{"LOWER": "digital"}, {"LOWER": "thread"}, {"LOWER": "engineer"}],[{"LOWER": "openresty"}],
            [{"LOWER": "idm"}],[{"LOWER": "oneidm"}],[{"LOWER": "datalake"}],[{"LOWER": "afrt"}],[{"LOWER": "aeip"}], [{"LOWER": "data lake"}])

Doc.set_extension('closure_note', default=None)
Doc.set_extension('desc', default=None)
Doc.set_extension('id', default=None)
Doc.set_extension('app', default=['unknown'])
Doc.set_extension('category', default=['general'])
Doc.set_extension('notes', default=[])
Doc.set_extension('attachments', default=[])
Doc.set_extension('highlight', default=None)

def custom_category_component(doc):
    scores = []
    for i in category_list:         
        s = doc.similarity(i[0])
        scores.append((i, s))
        orderedScores = sorted(scores, key=lambda tup: tup[1], reverse=True)
        topscore=orderedScores[0]
        catlist=[]
        for j in orderedScores:
            if float(j[1]) > 0.50:
                catlist.append(j[0].text)
        if len(catlist) != 0:
            doc._.category = catlist
    return doc

def application_component(doc):
    matches = matcher(doc)
    appNames = [doc[start:end].text for match_id, start, end in matches]
    for i in range(len(appNames)):         
        appNames[i] = appNames[i].lower() 
        if appMap.get(appNames[i]) != None:
            appNames[i] = appMap.get(appNames[i])
    if len(appNames) > 0:            
        doc._.app = list(set(appNames))
    return doc

nlp.add_pipe(application_component)
nlp.add_pipe(custom_category_component)


for doc, context in nlp.pipe(unprocessDocs, as_tuples=True):
    docIndex['id'] = docIndex['id'] + 1
    doc._.id = docIndex['id']
    doc._.desc = context['desc']
    doc._.closure_note = context['closure_note']
    # if len(context['noteArray']) > 0:
    #     doc._.notes = context['noteArray']
    sampleDocs.append(doc)

files = ['/data/DTE_DRB_General_Concepts.pdf','/data/DTE_DRB_Getting_Started.pdf']
for fileName in files:
    with open(fileName, 'rb') as f:
        doc = slate.PDF(f)
        for i, p in enumerate(doc):
            doc = nlp(p)
            doc._.desc = p
            docIndex['id'] = docIndex['id'] + 1
            doc._.id = docIndex['id']
            doc._.closure_note = "Page "+ str(i+1) +" of "+ fileName + " has more details on this"
            # doc._.highlight = para
            sampleDocs.append(doc)

    
def getSuggestions(desc):
    scores = []
    newDoc = nlp(desc)
    for doc in sampleDocs:
        #print(newDoc._.app, doc._.app, list(set(doc._.app).intersection(newDoc._.app)))
        #if ('unknown' in doc._.app) or (len(list(set(doc._.app).intersection(newDoc._.app))) > 0):
        is_intersect=0
        if ('unknown' not in doc._.app) and (len(list(set(doc._.app).intersection(newDoc._.app))) > 0):
            is_intersect=1
        scores.append({"intent": doc.text,"desc":doc._.desc, "app": doc._.app, "notes": doc._.notes,"attachments":doc._.attachments,
        "category": doc._.category, "score":doc.similarity(newDoc), "id":doc._.id,
        "closure_note": doc._.closure_note, "is_intersect": is_intersect})

    orderedScores = sorted(scores, key=lambda tup: tup["score"], reverse=True)

    return sorted(orderedScores[0:10], key=lambda tup: tup["is_intersect"], reverse=True)

def getDocById(id):
    for doc in sampleDocs:
        if doc._.id == id:
            return doc

    return None
def addNote(id, note):
    doc = getDocById(id)
    if doc != None:
        #doc._.notes.append(note)
        text = doc.text + ' ' + note
        newDoc = nlp(text)
        newDoc._.app = doc._.app
        newDoc._.desc = doc._.desc
        newDoc._.category = doc._.category
        newDoc._.id = doc._.id
        newDoc._.closure_note = doc._.closure_note 
        newDoc._.notes = doc._.notes
        newDoc._.attachments = doc._.attachments
        newDoc._.notes.append(note)
        # print(sampleDocs[636])
        # print(sampleDocs[0])
        # print(sampleDocs[635])
        sampleDocs.remove(doc)
        sampleDocs.insert(newDoc._.id-2,newDoc)
        
        # notes = ''
        # for note in newDoc._.notes:
        #     notes += note + ';'
        # df = pd.DataFrame([[doc._.desc, doc._.closure_note, notes]])
        # print(doc._.id, doc._.desc, doc._.closure_note, notes)
        # book = load_workbook('test.xlsx')
        # writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
        # writer.book = book
        # writer.sheets = {ws.title: ws for ws in book.worksheets}    
        # for sheetname in writer.sheets:
        #     df.to_excel(writer,sheet_name='Sheet1', startrow= doc._.id-1, index = False,header= False)
        # writer.save()    

def updateClosureNote(id, closure_note):
    doc = getDocById(id)
    if doc != None:
        doc._.closure_note = closure_note
        # df = pd.DataFrame([[doc._.desc, doc._.closure_note]])
        # book = load_workbook('test.xlsx')
        # writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
        # writer.book = book
        # writer.sheets = {ws.title: ws for ws in book.worksheets}    
        # for sheetname in writer.sheets:
        #     df.to_excel(writer,sheet_name='Sheet1', startrow= doc._.id-1, index = False,header= False)
        # writer.save()       

def addVotedResult(description,closurenote):
    doc = nlp(description)
    docIndex['id'] = docIndex['id'] + 1
    doc._.id = docIndex['id']
    doc._.desc = description
    doc._.closure_note = closurenote
    sampleDocs.append(doc)

def addFile(id, fileName):
    doc = getDocById(id)
    if doc != None:
        doc._.attachments.append(fileName)

def addTicket(issue, note):
    #doc._.notes.append(note)
    doc = nlp(issue)
    doc._.desc = issue
    docIndex['id'] = docIndex['id'] + 1
    doc._.id = docIndex['id']
    doc._.closure_note = note
    sampleDocs.append(doc)
    # create pd data frame
    # df = pd.DataFrame([[doc._.desc, doc._.closure_note]])
    # print(doc._.id, doc._.desc, doc._.closure_note)
    # book = load_workbook('test.xlsx')
    # writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    # writer.book = book
    # writer.sheets = {ws.title: ws for ws in book.worksheets}

    # for sheetname in writer.sheets:
    #     df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    # writer.save()

def removeTicket(id):
    doc = getDocById(id)
    if doc != None:
        sampleDocs.remove(doc)